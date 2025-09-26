from typing import List, Optional, Tuple
from fastapi import HTTPException, status
from bson import ObjectId
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io

from app.database.connection import db
from app.products.schemas import (
    ProductCreate, ProductUpdate, CategoryCreate, CategoryUpdate, ProductFilter
)
from app.products.models import Product, Category


class ProductService:

    @staticmethod
    async def _add_computed_fields(doc: dict) -> dict:
        """Add computed fields to product document"""
        # Convert ObjectId to string
        if "_id" in doc and not isinstance(doc["_id"], str):
            doc["_id"] = str(doc["_id"])
        
        # Handle backward compatibility for prices
        selling_price = doc.get("selling_price") or doc.get("price")
        doc["selling_price"] = selling_price
        
        # Add computed fields
        discount = doc.get("discount", 0)
        price = selling_price  # Use selling_price for calculations
        quantity = doc["quantity"]
        
        doc["final_price"] = price * (1 - (discount / 100)) if discount > 0 else price
        doc["is_low_stock"] = quantity < 10
        
        if quantity == 0:
            doc["stock_status"] = "out_of_stock"
        elif quantity < 10:
            doc["stock_status"] = "low_stock"
        elif quantity < 50:
            doc["stock_status"] = "medium_stock"
        else:
            doc["stock_status"] = "high_stock"
        
        # Add expiry date calculations
        expiry_date = doc.get("expiry_date")
        if expiry_date:
            now = datetime.utcnow()
            if isinstance(expiry_date, str):
                from dateutil.parser import parse
                expiry_date = parse(expiry_date)
            
            days_until_expiry = (expiry_date - now).days
            doc["days_until_expiry"] = days_until_expiry
            doc["is_expired"] = days_until_expiry < 0
        else:
            doc["days_until_expiry"] = None
            doc["is_expired"] = False
        
        # Add category name resolution
        category_id = doc.get("category_id")
        if category_id:
            try:
                category = await db.categories.find_one({"_id": ObjectId(category_id), "is_active": True})
                doc["category_name"] = category["name"] if category else "Unknown Category"
            except:
                doc["category_name"] = "Invalid Category"
        else:
            doc["category_name"] = None
            
        return doc

    ### CATEGORY METHODS

    @staticmethod
    async def create_category(category: CategoryCreate) -> Category:
        existing = await db.categories.find_one({"name": category.name})
        if existing:
            raise HTTPException(status_code=400, detail="Category already exists")
        category_data = category.dict()
        category_data.update({
            "is_active": True,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        })
        result = await db.categories.insert_one(category_data)
        category_data["_id"] = result.inserted_id
        category = Category(**category_data)
        return category

    @staticmethod
    async def get_categories(skip=0, limit=100) -> List[Category]:
        cursor = db.categories.find({"is_active": True}).skip(skip).limit(limit)
        return [Category(**cat) async for cat in cursor]

    @staticmethod
    async def get_category_by_id(category_id: str) -> Optional[Category]:
        category = await db.categories.find_one({"_id": ObjectId(category_id), "is_active": True})
        return Category(**category) if category else None

    @staticmethod
    async def update_category(category_id: str, category_update: CategoryUpdate) -> Optional[Category]:
        existing = await ProductService.get_category_by_id(category_id)
        if not existing:
            raise HTTPException(status_code=404, detail="Category not found")

        update_data = {k: v for k, v in category_update.dict().items() if v is not None}
        update_data["updated_at"] = datetime.utcnow()

        if "name" in update_data:
            if await db.categories.find_one({"name": update_data["name"], "_id": {"$ne": ObjectId(category_id)}}):
                raise HTTPException(status_code=400, detail="Category with this name already exists")

        await db.categories.update_one({"_id": ObjectId(category_id)}, {"$set": update_data})
        return await ProductService.get_category_by_id(category_id)

    @staticmethod
    async def delete_category(category_id: str) -> bool:
        category = await ProductService.get_category_by_id(category_id)
        if not category:
            raise HTTPException(status_code=404, detail="Category not found")

        products_count = await db.products.count_documents({"category_id": category_id, "is_active": True})
        if products_count > 0:
            raise HTTPException(status_code=400, detail="Cannot delete category with existing products")

        await db.categories.update_one({"_id": ObjectId(category_id)}, {"$set": {"is_active": False}})
        return True

    @staticmethod
    async def initialize_default_categories() -> List[Category]:
        count = await db.categories.count_documents({})
        if count > 0:
            cursor = db.categories.find()
            return [Category(**cat) async for cat in cursor]

        default_categories = [
            {"name": "إلكترونيات", "description": "الأجهزة الإلكترونية والإكسسوارات"},
            {"name": "بقالة", "description": "الطعام والمشروبات"},
            {"name": "ملابس", "description": "الملابس والإكسسوارات"},
            {"name": "منزل وحديقة", "description": "تحسين المنزل ومستلزمات الحديقة"},
        ]

        created = []
        for cat in default_categories:
            cat.update({
                "is_active": True,
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            })
            result = await db.categories.insert_one(cat)
            cat["_id"] = result.inserted_id
            created.append(Category(**cat))
        return created

    ### PRODUCT METHODS

    @staticmethod
    async def create_product(product: ProductCreate) -> dict:
        if not await ProductService.get_category_by_id(product.category_id):
            raise HTTPException(status_code=404, detail="Category not found")

        if product.sku:
            if await db.products.find_one({"sku": product.sku}):
                raise HTTPException(status_code=400, detail="SKU already exists")

        if product.product_id:
            if await db.products.find_one({"product_id": product.product_id}):
                raise HTTPException(status_code=400, detail="Product ID already exists")
        else:
            # Generate product_id if not provided
            product.product_id = f"PRD-{ObjectId()}"

        data = product.dict()
        data.update({
            "is_active": True,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        })
        result = await db.products.insert_one(data)
        data["_id"] = result.inserted_id
        
        return await ProductService._add_computed_fields(data)

    @staticmethod
    async def get_product_by_id(product_id: str) -> Optional[dict]:
        doc = await db.products.find_one({"_id": ObjectId(product_id), "is_active": True})
        return await ProductService._add_computed_fields(doc) if doc else None

    @staticmethod
    async def get_product_by_product_id(product_id: str) -> Optional[dict]:
        doc = await db.products.find_one({"product_id": product_id, "is_active": True})
        return await ProductService._add_computed_fields(doc) if doc else None

    @staticmethod
    async def get_products(skip=0, limit=100, filters: Optional[ProductFilter] = None) -> Tuple[List[dict], int]:
        query = {"is_active": True}

        if filters:
            if filters.category_id:
                query["category_id"] = filters.category_id
            if filters.search:
                query["$or"] = [
                    {"name": {"$regex": filters.search, "$options": "i"}},
                    {"description": {"$regex": filters.search, "$options": "i"}},
                    {"sku": {"$regex": filters.search, "$options": "i"}},
                    {"product_id": {"$regex": filters.search, "$options": "i"}},
                ]
            if filters.min_price is not None:
                query["price"] = {"$gte": filters.min_price}
            if filters.max_price is not None:
                query.setdefault("price", {}).update({"$lte": filters.max_price})
            if filters.in_stock_only:
                query["quantity"] = {"$gt": 0}
            if filters.low_stock_only:
                query["quantity"] = {"$lt": 10}

        cursor = db.products.find(query).skip(skip).limit(limit)
        products = []
        async for doc in cursor:
            products.append(await ProductService._add_computed_fields(doc))
            
        total = await db.products.count_documents(query)
        return products, total

    @staticmethod
    async def update_product(product_id: str, update: ProductUpdate) -> dict:
        product = await ProductService.get_product_by_id(product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        if update.category_id:
            if not await ProductService.get_category_by_id(update.category_id):
                raise HTTPException(status_code=404, detail="Category not found")

        if update.sku and update.sku != product["sku"]:
            if await db.products.find_one({"sku": update.sku, "_id": {"$ne": ObjectId(product_id)}}):
                raise HTTPException(status_code=400, detail="SKU already exists")

        if update.product_id and update.product_id != product["product_id"]:
            if await db.products.find_one({"product_id": update.product_id, "_id": {"$ne": ObjectId(product_id)}}):
                raise HTTPException(status_code=400, detail="Product ID already exists")

        update_data = {k: v for k, v in update.dict().items() if v is not None}
        update_data["updated_at"] = datetime.utcnow()

        await db.products.update_one({"_id": ObjectId(product_id)}, {"$set": update_data})
        return await ProductService.get_product_by_id(product_id)

    @staticmethod
    async def delete_product(product_id: str) -> bool:
        product = await ProductService.get_product_by_id(product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        await db.products.update_one({"_id": ObjectId(product_id)}, {"$set": {"is_active": False}})
        return True

    @staticmethod
    async def update_stock(product_id: str, quantity: int) -> dict:
        product = await ProductService.get_product_by_id(product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        await db.products.update_one({"_id": ObjectId(product_id)}, {"$set": {"quantity": quantity}})
        return await ProductService.get_product_by_id(product_id)

    @staticmethod
    async def get_low_stock_products(threshold: int = 10) -> List[dict]:
        cursor = db.products.find({"is_active": True, "quantity": {"$lt": threshold}})
        products = []
        async for doc in cursor:
            products.append(await ProductService._add_computed_fields(doc))
            
        return products

    @staticmethod
    async def get_expired_products() -> List[dict]:
        """Get all products that have expired or are expiring soon (within 7 days)"""
        now = datetime.utcnow()
        cursor = db.products.find({
            "is_active": True, 
            "expiry_date": {"$lte": now}
        })
        products = []
        async for doc in cursor:
            products.append(await ProductService._add_computed_fields(doc))
            
        return products

    @staticmethod
    async def get_expiring_soon_products(days: int = 7) -> List[dict]:
        """Get products expiring within specified days"""
        now = datetime.utcnow()
        from datetime import timedelta
        expiry_threshold = now + timedelta(days=days)
        
        cursor = db.products.find({
            "is_active": True, 
            "expiry_date": {"$gte": now, "$lte": expiry_threshold}
        })
        products = []
        async for doc in cursor:
            products.append(await ProductService._add_computed_fields(doc))
            
        return products
    
    @staticmethod
    async def create_inventory_excel(products: List[dict]) -> bytes:
        """Create Excel file with inventory data"""
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "جرد المنتجات"
        
        # Define headers
        headers = [
            "ID المنتج",
            "اسم المنتج", 
            "الوحدة",
            "كمية المنتج",
            "سعر البيع",
            "سعر الشراء", 
            "إجمالي السعر"
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row, product in enumerate(products, 2):
            # Product ID
            ws.cell(row=row, column=1, value=product.get("product_id", "N/A"))
            
            # Product Name
            ws.cell(row=row, column=2, value=product.get("name", ""))
            
            # Unit (قطعة or دسته)
            unit_text = "دسته" if product.get("size_unit") == "dozen" else "قطعة"
            ws.cell(row=row, column=3, value=unit_text)
            
            # Quantity
            quantity = product.get("quantity", 0)
            ws.cell(row=row, column=4, value=quantity)
            
            # Selling Price
            selling_price = product.get("selling_price") or product.get("price", 0)
            ws.cell(row=row, column=5, value=selling_price)
            
            # Buying Price
            buying_price = product.get("buying_price", 0)
            ws.cell(row=row, column=6, value=buying_price if buying_price else "غير محدد")
            
            # Total Price (Buying Price * Quantity)
            if buying_price:
                total_price = buying_price * quantity
                ws.cell(row=row, column=7, value=total_price)
            else:
                ws.cell(row=row, column=7, value="غير محدد")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add totals row
        total_row = len(products) + 3
        ws.cell(row=total_row, column=1, value="الإجمالي:").font = Font(bold=True)
        
        # Calculate total inventory value
        total_value = 0
        total_quantity = 0
        
        for product in products:
            quantity = product.get("quantity", 0)
            buying_price = product.get("buying_price")
            total_quantity += quantity
            if buying_price:
                total_value += buying_price * quantity
        
        ws.cell(row=total_row, column=4, value=total_quantity).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=f"{total_value:.2f} جنيه").font = Font(bold=True)
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
